from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="FF1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="FF70AD47")
SUMMARY_LABEL_FILL = PatternFill("solid", fgColor="FFD9EAF7")
SUMMARY_VALUE_FILL = PatternFill("solid", fgColor="FFEAF3F8")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True, color="FFFFFF")
SUMMARY_LABEL_FONT = Font(bold=True, color="FF1F4E78")
SUMMARY_VALUE_FONT = Font(bold=True, color="FF000000")
BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="FF808080"),
    right=Side(style="thin", color="FF808080"),
    top=Side(style="thin", color="FF808080"),
    bottom=Side(style="thin", color="FF808080"),
)
AMOUNT_NUMBER_FORMAT = '#,##0.00;[Red]-#,##0.00'
DATE_NUMBER_FORMAT = "yyyy-mm-dd"
AMOUNT_KEYWORDS = (
    "金額",
    "金额",
    "售價",
    "售价",
    "售賣價",
    "售卖价",
    "價格",
    "价格",
    "費用",
    "费用",
    "總和",
    "总和",
    "合計",
    "合计",
    "平均值",
)
DATE_KEYWORDS = ("日期", "日付", "date", "時間", "时间", "time")
SECTION_TITLES = {"總覽", "分組統計", "月份統計", "Top10排行"}
SUMMARY_TITLE = "摘要"


def format_workbook(workbook_path: str | Path) -> str:
    """Apply v2.1 presentation styles to an exported workbook."""
    path = Path(workbook_path)
    workbook = load_workbook(path)

    if "統計報表" in workbook.sheetnames:
        _prepare_statistics_sheet(workbook["統計報表"])

    for worksheet in workbook.worksheets:
        _format_sheet(worksheet)

    workbook.save(path)
    return str(path)


def _format_sheet(worksheet) -> None:
    if worksheet.max_row == 0 or worksheet.max_column == 0:
        return

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_rows = set(_detect_header_rows(worksheet))
    amount_columns = set(_detect_amount_columns(worksheet, header_rows))
    date_columns = set(_detect_date_columns(worksheet, header_rows))

    for row in worksheet.iter_rows():
        row_index = row[0].row
        is_section_row = _is_section_row(worksheet, row_index)
        is_summary_title_row = _is_summary_title_row(worksheet, row_index)
        is_summary_value_row = _is_summary_value_row(worksheet, row_index)
        is_header_row = row_index in header_rows

        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = BODY_ALIGNMENT

            if is_summary_title_row:
                cell.font = SECTION_FONT
                cell.fill = SECTION_FILL
                cell.alignment = HEADER_ALIGNMENT
            elif is_summary_value_row:
                _format_summary_cell(worksheet, cell)
            elif is_section_row:
                cell.font = SECTION_FONT
                cell.fill = SECTION_FILL
                cell.alignment = HEADER_ALIGNMENT
            elif is_header_row:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGNMENT
            elif _is_date_value_cell(cell, date_columns):
                cell.number_format = DATE_NUMBER_FORMAT
            elif _is_amount_value_cell(worksheet, cell, amount_columns):
                cell.number_format = AMOUNT_NUMBER_FORMAT

    _auto_adjust_column_widths(worksheet)


def _prepare_statistics_sheet(worksheet) -> None:
    _insert_summary_block(worksheet)
    _add_top10_ranking(worksheet)


def _insert_summary_block(worksheet) -> None:
    if worksheet.cell(row=1, column=1).value == SUMMARY_TITLE:
        return

    summary_values = _extract_overview_values(worksheet)
    if not summary_values:
        return

    worksheet.insert_rows(1, amount=3)
    worksheet.cell(row=1, column=1, value=SUMMARY_TITLE)

    labels_and_values = [
        ("總筆數", summary_values.get("總筆數")),
        ("金額總和", summary_values.get("總和")),
        ("金額平均值", summary_values.get("平均值")),
    ]
    column = 1
    for label, value in labels_and_values:
        if value is None:
            continue
        worksheet.cell(row=2, column=column, value=label)
        worksheet.cell(row=2, column=column + 1, value=value)
        column += 2


def _extract_overview_values(worksheet) -> dict[str, object]:
    overview_row = _find_section_row(worksheet, "總覽")
    if overview_row is None:
        return {}

    values: dict[str, object] = {}
    row_index = overview_row + 2
    while row_index <= worksheet.max_row:
        label = worksheet.cell(row=row_index, column=1).value
        value = worksheet.cell(row=row_index, column=2).value

        if label in SECTION_TITLES:
            break
        if label in (None, ""):
            row_index += 1
            continue

        label_text = str(label)
        if "總筆數" in label_text:
            values["總筆數"] = value
        elif "總和" in label_text:
            values["總和"] = value
        elif "平均值" in label_text:
            values["平均值"] = value

        row_index += 1

    return values


def _add_top10_ranking(worksheet) -> None:
    section_row = _find_section_row(worksheet, "Top10排行")
    if section_row is None:
        return

    header_row = section_row + 1
    if worksheet.cell(row=header_row, column=1).value == "排名":
        return

    end_row = _find_section_end_row(worksheet, header_row)
    if end_row < header_row:
        return

    max_column = worksheet.max_column
    for row_index in range(header_row, end_row + 1):
        for column_index in range(max_column, 0, -1):
            source = worksheet.cell(row=row_index, column=column_index)
            target = worksheet.cell(row=row_index, column=column_index + 1)
            _copy_cell(source, target)
            _clear_cell(source)

    worksheet.cell(row=header_row, column=1, value="排名")
    for rank, row_index in enumerate(range(header_row + 1, end_row + 1), start=1):
        worksheet.cell(row=row_index, column=1, value=rank)


def _find_section_row(worksheet, title: str) -> int | None:
    for row_index in range(1, worksheet.max_row + 1):
        if worksheet.cell(row=row_index, column=1).value == title:
            return row_index
    return None


def _find_section_end_row(worksheet, start_row: int) -> int:
    row_index = start_row
    while row_index <= worksheet.max_row:
        first_value = worksheet.cell(row=row_index, column=1).value
        if row_index > start_row and first_value in SECTION_TITLES:
            return row_index - 1
        if row_index > start_row and _non_empty_count(worksheet, row_index) == 0:
            return row_index - 1
        row_index += 1
    return worksheet.max_row


def _copy_cell(source, target) -> None:
    target.value = source.value
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format


def _clear_cell(cell) -> None:
    cell.value = None


def _detect_header_rows(worksheet) -> Iterable[int]:
    for row_index in range(1, worksheet.max_row + 1):
        if _is_section_row(worksheet, row_index):
            yield row_index
            next_row = row_index + 1
            if next_row <= worksheet.max_row and _non_empty_count(worksheet, next_row) >= 2:
                yield next_row
            continue

        if row_index == 1 and _non_empty_count(worksheet, row_index) >= 2:
            yield row_index


def _detect_amount_columns(worksheet, header_rows: set[int]) -> Iterable[int]:
    for row_index in header_rows:
        for cell in worksheet[row_index]:
            if cell.value is None:
                continue
            if _is_amount_header(str(cell.value)):
                yield cell.column


def _detect_date_columns(worksheet, header_rows: set[int]) -> Iterable[int]:
    for row_index in header_rows:
        for cell in worksheet[row_index]:
            if cell.value is None:
                continue
            if _is_date_header(str(cell.value)):
                yield cell.column


def _is_amount_header(header: str) -> bool:
    normalized = header.strip().lower()
    return any(keyword.lower() in normalized for keyword in AMOUNT_KEYWORDS)


def _is_date_header(header: str) -> bool:
    normalized = header.strip().lower()
    return any(keyword.lower() in normalized for keyword in DATE_KEYWORDS)


def _is_amount_value_cell(worksheet, cell, amount_columns: set[int]) -> bool:
    if not isinstance(cell.value, (int, float)):
        return False

    if cell.column in amount_columns:
        return True

    left_value = worksheet.cell(row=cell.row, column=max(cell.column - 1, 1)).value
    return left_value is not None and _is_amount_header(str(left_value))


def _is_date_value_cell(cell, date_columns: set[int]) -> bool:
    if cell.column not in date_columns or cell.value in (None, ""):
        return False

    return isinstance(cell.value, (int, float, datetime))


def _is_section_row(worksheet, row_index: int) -> bool:
    first_value = worksheet.cell(row=row_index, column=1).value
    if first_value not in SECTION_TITLES:
        return False
    return _non_empty_count(worksheet, row_index) == 1


def _is_summary_title_row(worksheet, row_index: int) -> bool:
    return worksheet.cell(row=row_index, column=1).value == SUMMARY_TITLE


def _is_summary_value_row(worksheet, row_index: int) -> bool:
    return (
        row_index == 2
        and worksheet.cell(row=1, column=1).value == SUMMARY_TITLE
    )


def _format_summary_cell(worksheet, cell) -> None:
    cell.alignment = HEADER_ALIGNMENT
    if cell.column % 2 == 1:
        cell.font = SUMMARY_LABEL_FONT
        cell.fill = SUMMARY_LABEL_FILL
        return

    cell.font = SUMMARY_VALUE_FONT
    cell.fill = SUMMARY_VALUE_FILL
    left_value = worksheet.cell(row=cell.row, column=cell.column - 1).value
    if left_value is not None and _is_amount_header(str(left_value)) and isinstance(cell.value, (int, float)):
        cell.number_format = AMOUNT_NUMBER_FORMAT


def _non_empty_count(worksheet, row_index: int) -> int:
    return sum(
        1
        for cell in worksheet[row_index]
        if cell.value not in (None, "")
    )


def _auto_adjust_column_widths(worksheet) -> None:
    for column_index in range(1, worksheet.max_column + 1):
        max_width = 0
        for row_index in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row=row_index, column=column_index).value
            if value is None:
                continue
            max_width = max(max_width, _display_width(value))

        width = min(max(max_width + 2, 12), 48)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def _display_width(value: object) -> int:
    text = str(value)
    width = 0
    for char in text:
        width += 2 if ord(char) > 127 else 1
    return width
