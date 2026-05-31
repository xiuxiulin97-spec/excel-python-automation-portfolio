from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from exporter import export_customer_report


def test_export_customer_report_creates_timestamped_workbook_with_three_sheets(tmp_path: Path):
    original_df = pd.DataFrame([{"姓名": "王小明", "電話": "0912345678", "Email": "test@example.com", "公司": "A公司"}])
    cleaned_df = original_df.copy()
    issues_df = pd.DataFrame(columns=["姓名", "電話", "Email", "公司", "問題"])

    output_path = export_customer_report(original_df, cleaned_df, issues_df, output_dir=str(tmp_path))

    assert Path(output_path).exists()
    assert Path(output_path).name.startswith("cleaned_customers_")
    assert Path(output_path).suffix == ".xlsx"

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["原始資料", "清理後資料", "問題資料"]
