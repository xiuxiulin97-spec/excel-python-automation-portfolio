from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from report import build_summary, export_report, validate_columns


def test_build_summary_groups_amount_by_customer():
    df = pd.DataFrame(
        [
            {"客戶": "A 公司", "金額": 100},
            {"客戶": "B 公司", "金額": 200},
            {"客戶": "A 公司", "金額": 300},
        ]
    )

    summary = build_summary(df)

    assert summary.to_dict("records") == [
        {"客戶": "A 公司", "金額總和": 400},
        {"客戶": "B 公司", "金額總和": 200},
    ]


def test_validate_columns_requires_customer_column():
    df = pd.DataFrame([{"金額": 100}])

    with pytest.raises(ValueError, match="客戶"):
        validate_columns(df)


def test_validate_columns_requires_amount_column():
    df = pd.DataFrame([{"客戶": "A 公司"}])

    with pytest.raises(ValueError, match="金額"):
        validate_columns(df)


def test_build_summary_rejects_invalid_amount_values():
    df = pd.DataFrame([{"客戶": "A 公司", "金額": "不是數字"}])

    with pytest.raises(ValueError, match="金額"):
        build_summary(df)


def test_export_report_creates_workbook_with_three_sheets(tmp_path: Path):
    original_df = pd.DataFrame(
        [
            {"客戶": "A 公司", "金額": 100},
            {"客戶": "A 公司", "金額": 300},
            {"客戶": "B 公司", "金額": 200},
        ]
    )
    cleaned_df = original_df.copy()

    output_path = export_report(original_df, cleaned_df, output_dir=str(tmp_path))

    assert Path(output_path).exists()
    assert Path(output_path).name.startswith("result_")
    assert Path(output_path).suffix == ".xlsx"

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["原始資料", "清理後資料", "統計報表"]
