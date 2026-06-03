from __future__ import annotations

from pathlib import Path
from datetime import datetime

from openpyxl import Workbook, load_workbook

from excel_formatter import format_workbook
from ui import export_dynamic_report


def test_format_workbook_applies_basic_report_styles(tmp_path: Path):
    workbook_path = tmp_path / "report.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "清理後資料"
    sheet.append(["客戶", "金額", "備註"])
    sheet.append(["A公司", 1200, "首筆"])
    sheet.append(["B公司", 30500.5, "第二筆"])

    stats_sheet = workbook.create_sheet("統計報表")
    stats_sheet.append(["總覽"])
    stats_sheet.append(["項目", "值"])
    stats_sheet.append(["金額總和", 1200])
    stats_sheet.append([])
    stats_sheet.append(["分組統計"])
    stats_sheet.append(["客戶", "筆數", "金額總和"])
    stats_sheet.append(["A公司", 1, 1200])

    workbook.save(workbook_path)

    format_workbook(workbook_path)

    formatted = load_workbook(workbook_path)
    formatted_sheet = formatted["清理後資料"]
    formatted_stats_sheet = formatted["統計報表"]

    assert formatted_sheet.freeze_panes == "A2"
    assert formatted_sheet.auto_filter.ref == "A1:C3"
    assert formatted_sheet["A1"].font.bold is True
    assert formatted_sheet["A1"].fill.fgColor.rgb == "FF1F4E78"
    assert formatted_sheet["A1"].border.bottom.style == "thin"
    assert formatted_sheet["A1"].border.bottom.color.rgb == "FF808080"
    assert formatted_sheet["A2"].border.bottom.style == "thin"
    assert formatted_sheet["A2"].border.bottom.color.rgb == "FF808080"
    assert formatted_sheet["B2"].number_format == '#,##0.00;[Red]-#,##0.00'
    assert formatted_sheet.column_dimensions["B"].width > 10

    assert formatted_stats_sheet.freeze_panes == "A2"
    assert formatted_stats_sheet["A1"].font.bold is True
    overview_total_row = _find_row(formatted_stats_sheet, "金額總和")
    group_header_row = _find_row(formatted_stats_sheet, "分組統計") + 1
    group_data_row = group_header_row + 1

    assert formatted_stats_sheet.cell(overview_total_row, 2).number_format == '#,##0.00;[Red]-#,##0.00'
    assert formatted_stats_sheet.cell(group_data_row, 3).number_format == '#,##0.00;[Red]-#,##0.00'


def test_format_workbook_applies_second_batch_report_styles(tmp_path: Path):
    workbook_path = tmp_path / "report.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "清理後資料"
    sheet.append(["日期", "客戶", "金額"])
    sheet.append([datetime(2026, 6, 3), "A公司", 1200])
    sheet.append([45686, "B公司", 30500.5])

    stats_sheet = workbook.create_sheet("統計報表")
    stats_sheet.append(["總覽"])
    stats_sheet.append(["項目", "值"])
    stats_sheet.append(["總筆數", 2])
    stats_sheet.append(["金額總和", 31700.5])
    stats_sheet.append(["金額平均值", 15850.25])
    stats_sheet.append([])
    stats_sheet.append(["Top10排行"])
    stats_sheet.append(["客戶", "金額"])
    stats_sheet.append(["B公司", 30500.5])
    stats_sheet.append(["A公司", 1200])
    workbook.save(workbook_path)

    format_workbook(workbook_path)

    formatted = load_workbook(workbook_path)
    formatted_sheet = formatted["清理後資料"]
    formatted_stats_sheet = formatted["統計報表"]

    assert formatted_sheet["A2"].number_format == "yyyy-mm-dd"
    assert formatted_sheet["A3"].number_format == "yyyy-mm-dd"

    assert formatted_stats_sheet["A1"].value == "摘要"
    assert formatted_stats_sheet["A2"].value == "總筆數"
    assert formatted_stats_sheet["B2"].value == 2
    assert formatted_stats_sheet["C2"].value == "金額總和"
    assert formatted_stats_sheet["D2"].value == 31700.5
    assert formatted_stats_sheet["D2"].number_format == '#,##0.00;[Red]-#,##0.00'

    top10_header_row = _find_row(formatted_stats_sheet, "Top10排行") + 1
    assert formatted_stats_sheet.cell(top10_header_row, 1).value == "排名"
    assert formatted_stats_sheet.cell(top10_header_row + 1, 1).value == 1
    assert formatted_stats_sheet.cell(top10_header_row + 2, 1).value == 2

    for worksheet in formatted.worksheets:
        assert worksheet.sheet_view.showGridLines is False
        assert worksheet.freeze_panes == "A2"
        assert worksheet["A1"].border.bottom.color.rgb == "FF808080"


def test_export_dynamic_report_formats_generated_workbook(tmp_path: Path):
    import pandas as pd

    original_df = pd.DataFrame(
        [
            {"客戶": "A公司", "金額": 1200},
            {"客戶": "B公司", "金額": 30500.5},
        ]
    )

    output_path = export_dynamic_report(
        original_df=original_df,
        cleaned_df=original_df,
        group_column="客戶",
        value_column="金額",
        output_dir=str(tmp_path),
    )

    workbook = load_workbook(output_path)
    cleaned_sheet = workbook["清理後資料"]

    assert cleaned_sheet.freeze_panes == "A2"
    assert cleaned_sheet.auto_filter.ref == "A1:B3"
    assert cleaned_sheet["A1"].font.bold is True
    assert cleaned_sheet["B2"].number_format == '#,##0.00;[Red]-#,##0.00'


def _find_row(worksheet, value: str) -> int:
    for row in range(1, worksheet.max_row + 1):
        if worksheet.cell(row, 1).value == value:
            return row
    raise AssertionError(f"找不到列：{value}")
